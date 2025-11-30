#!/usr/bin/env python3
"""
Parse WooCommerce SnappPay Orders Data

This script parses the JSON file containing WooCommerce orders with SnappPay
payment information and outputs a formatted file with:
- First Name
- Last Name
- Order ID
- SnappPay Token
- Transaction ID
"""

import json
import csv
import sys
from pathlib import Path
from datetime import datetime

# Try to import optional libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def parse_orders(input_file, output_format='csv'):
    """
    Parse orders from JSON file and extract required information.

    Args:
        input_file: Path to input JSON file
        output_format: Output format ('csv', 'json', 'txt')

    Returns:
        List of parsed order dictionaries
    """
    # Read input JSON file
    try:
        with open(input_file, 'r', encoding='utf-8') as f:
            orders = json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: File not found: {input_file}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON file: {e}")
        sys.exit(1)

    print(f"📊 Found {len(orders)} orders in input file")

    # Parse and extract required fields
    parsed_orders = []
    for order in orders:
        parsed_order = {
            'first_name': order.get('billing', {}).get('first_name', '') or
                         order.get('shipping', {}).get('first_name', ''),
            'last_name': order.get('billing', {}).get('last_name', '') or
                        order.get('shipping', {}).get('last_name', ''),
            'order_id': order.get('orderId', ''),
            'snapp_pay_token': order.get('snappPayToken', ''),
            'transaction_id': order.get('transactionId', '') or '',
        }
        parsed_orders.append(parsed_order)

    print(f"✅ Parsed {len(parsed_orders)} orders")
    return parsed_orders

def write_csv(output_file, orders):
    """Write orders to CSV file"""
    if not orders:
        print("⚠️  No orders to write")
        return

    fieldnames = ['first_name', 'last_name', 'order_id', 'snapp_pay_token', 'transaction_id']

    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(orders)

    print(f"✅ Written {len(orders)} orders to CSV: {output_file}")

def write_json(output_file, orders):
    """Write orders to JSON file"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(orders, f, ensure_ascii=False, indent=2)

    print(f"✅ Written {len(orders)} orders to JSON: {output_file}")

def write_txt(output_file, orders):
    """Write orders to formatted text file"""
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("WooCommerce SnappPay Orders - Parsed Data\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")

        for i, order in enumerate(orders, 1):
            f.write(f"Order #{i}\n")
            f.write(f"  First Name: {order['first_name']}\n")
            f.write(f"  Last Name: {order['last_name']}\n")
            f.write(f"  Order ID: {order['order_id']}\n")
            f.write(f"  SnappPay Token: {order['snapp_pay_token']}\n")
            f.write(f"  Transaction ID: {order['transaction_id']}\n")
            f.write("-" * 80 + "\n\n")

    print(f"✅ Written {len(orders)} orders to text file: {output_file}")

def write_xlsx(output_file, orders):
    """Write orders to Excel XLSX file with beautiful styling"""
    if not HAS_OPENPYXL:
        print("❌ Error: openpyxl library not installed")
        print("   Install it with: pip install openpyxl")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "SnappPay Orders"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column headers
    headers = ['First Name', 'Last Name', 'Order ID', 'SnappPay Token', 'Transaction ID']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write data
    for row_num, order in enumerate(orders, 2):
        ws.cell(row=row_num, column=1, value=order['first_name']).border = border_style
        ws.cell(row=row_num, column=2, value=order['last_name']).border = border_style
        ws.cell(row=row_num, column=3, value=order['order_id']).border = border_style
        ws.cell(row=row_num, column=4, value=order['snapp_pay_token']).border = border_style
        ws.cell(row=row_num, column=5, value=order['transaction_id'] or '').border = border_style

        # Alternate row colors
        if row_num % 2 == 0:
            fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = fill

    # Auto-adjust column widths
    column_widths = [20, 20, 12, 40, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 30

    # Save file
    wb.save(output_file)
    print(f"✅ Written {len(orders)} orders to Excel file: {output_file}")

def write_html(output_file, orders):
    """Write orders to beautiful HTML file"""
    html_content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SnappPay Orders - Parsed Data</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}

        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}

        .stats {{
            display: flex;
            justify-content: space-around;
            padding: 20px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}

        .stat-item {{
            text-align: center;
        }}

        .stat-value {{
            font-size: 32px;
            font-weight: bold;
            color: #667eea;
        }}

        .stat-label {{
            font-size: 14px;
            color: #6c757d;
            margin-top: 5px;
        }}

        .table-wrapper {{
            overflow-x: auto;
            padding: 20px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}

        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}

        th {{
            padding: 15px;
            text-align: right;
            font-weight: 600;
            border: 1px solid rgba(255,255,255,0.2);
        }}

        td {{
            padding: 12px 15px;
            border: 1px solid #e9ecef;
        }}

        tbody tr {{
            transition: background-color 0.2s;
        }}

        tbody tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}

        tbody tr:hover {{
            background-color: #e7f3ff;
            cursor: pointer;
        }}

        .order-id {{
            font-weight: 600;
            color: #667eea;
        }}

        .token {{
            font-family: 'Courier New', monospace;
            font-size: 12px;
            color: #495057;
            word-break: break-all;
        }}

        .transaction-id {{
            font-family: 'Courier New', monospace;
            font-size: 12px;
            color: #28a745;
        }}

        .footer {{
            padding: 20px;
            text-align: center;
            color: #6c757d;
            font-size: 12px;
            border-top: 2px solid #e9ecef;
            background: #f8f9fa;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 SnappPay Orders - Parsed Data</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>

        <div class="stats">
            <div class="stat-item">
                <div class="stat-value">{len(orders)}</div>
                <div class="stat-label">Total Orders</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(1 for o in orders if o['snapp_pay_token'])}</div>
                <div class="stat-label">With Tokens</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(1 for o in orders if o['transaction_id'])}</div>
                <div class="stat-label">With Transaction IDs</div>
            </div>
        </div>

        <div class="table-wrapper">
            <table>
                <thead>
                    <tr>
                        <th>First Name</th>
                        <th>Last Name</th>
                        <th>Order ID</th>
                        <th>SnappPay Token</th>
                        <th>Transaction ID</th>
                    </tr>
                </thead>
                <tbody>
"""

    for order in orders:
        html_content += f"""
                    <tr>
                        <td>{order['first_name']}</td>
                        <td>{order['last_name']}</td>
                        <td class="order-id">{order['order_id']}</td>
                        <td class="token">{order['snapp_pay_token']}</td>
                        <td class="transaction-id">{order['transaction_id'] or ''}</td>
                    </tr>
"""

    html_content += """
                </tbody>
            </table>
        </div>

        <div class="footer">
            <p>SnappPay Orders Data - Infinity Store</p>
        </div>
    </div>
</body>
</html>
"""

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"✅ Written {len(orders)} orders to HTML file: {output_file}")

def main():
    # Determine input file path
    script_dir = Path(__file__).parent
    default_input = script_dir / 'woocommerce-guest-orders-snapppay-data.json'

    # Get input file from command line or use default
    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
    else:
        input_file = default_input

    # Check if file exists
    if not input_file.exists():
        print(f"❌ Error: Input file not found: {input_file}")
        print(f"   Usage: python {Path(__file__).name} [input_file.json]")
        sys.exit(1)

    # Determine output format (default: xlsx)
    output_format = 'xlsx'
    if len(sys.argv) > 2:
        output_format = sys.argv[2].lower()
        if output_format not in ['csv', 'json', 'txt', 'xlsx', 'html']:
            print(f"⚠️  Invalid output format '{output_format}', using 'xlsx'")
            output_format = 'xlsx'

    print("🚀 Starting SnappPay Orders Parser")
    print("=" * 60)
    print(f"📂 Input file: {input_file}")
    print(f"📝 Output format: {output_format.upper()}")
    print()

    # Parse orders
    orders = parse_orders(input_file)

    if not orders:
        print("⚠️  No orders found to process")
        sys.exit(0)

    # Generate output filename
    output_file = input_file.parent / f"snapppay-orders-parsed.{output_format}"

    # Write output based on format
    if output_format == 'csv':
        write_csv(output_file, orders)
    elif output_format == 'json':
        write_json(output_file, orders)
    elif output_format == 'txt':
        write_txt(output_file, orders)
    elif output_format == 'xlsx':
        write_xlsx(output_file, orders)
    elif output_format == 'html':
        write_html(output_file, orders)

    # Print summary
    print()
    print("📊 Summary:")
    print(f"   Total orders: {len(orders)}")
    print(f"   Orders with tokens: {sum(1 for o in orders if o['snapp_pay_token'])}")
    print(f"   Orders with transaction IDs: {sum(1 for o in orders if o['transaction_id'])}")
    print(f"   Output file: {output_file}")

    # Print sample
    if orders:
        print()
        print("📋 Sample output (first order):")
        sample = orders[0]
        print(f"   First Name: {sample['first_name']}")
        print(f"   Last Name: {sample['last_name']}")
        print(f"   Order ID: {sample['order_id']}")
        print(f"   SnappPay Token: {sample['snapp_pay_token']}")
        print(f"   Transaction ID: {sample['transaction_id']}")

if __name__ == '__main__':
    main()

