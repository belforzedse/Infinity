#!/usr/bin/env python3
"""
Extract WooCommerce SnappPay Orders and Match Phone Numbers

This script:
1. Extracts SnappPay orders without phone numbers from WooCommerce
2. Searches for orders with matching user names that have phone numbers
3. Creates a guessed orders result matching names to phone numbers
"""

import json
import sys
import time
import base64
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import requests
from requests.auth import HTTPBasicAuth

# Try to import optional libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

class WooCommerceClient:
    """WooCommerce REST API Client"""

    def __init__(self, base_url: str, consumer_key: str, consumer_secret: str):
        self.base_url = base_url.rstrip('/')
        self.auth = HTTPBasicAuth(consumer_key, consumer_secret)
        self.session = requests.Session()
        self.session.auth = self.auth

    def get_orders(self, params: Dict) -> tuple:
        """
        Fetch orders from WooCommerce API

        Returns:
            (orders_list, total_pages, total_items)
        """
        url = f"{self.base_url}/wp-json/wc/v3/orders"

        # Add auth as query params (some servers block Basic Auth headers)
        params = params.copy()
        params['consumer_key'] = self.auth.username
        params['consumer_secret'] = self.auth.password

        try:
            response = self.session.get(url, params=params, timeout=60)
            response.raise_for_status()

            orders = response.json()
            total_pages = int(response.headers.get('X-WP-TotalPages', 1))
            total_items = int(response.headers.get('X-WP-Total', 0))

            return orders, total_pages, total_items
        except requests.exceptions.RequestException as e:
            print(f"❌ API Error: {e}")
            if hasattr(e.response, 'text'):
                print(f"   Response: {e.response.text[:200]}")
            raise

def extract_snapppay_token(order: Dict) -> Optional[str]:
    """Extract SnappPay token from order meta_data or stored JSON value"""
    # Check if we have stored value from JSON
    if '_snapppay_token' in order:
        return order.get('_snapppay_token')

    # Otherwise check meta_data
    meta_data = order.get('meta_data', [])
    if not meta_data:
        return None

    for meta in meta_data:
        key = meta.get('key', '')
        if key in ['_order_spp_token', '_paymentToken']:
            return meta.get('value')
    return None

def extract_transaction_id(order: Dict) -> Optional[str]:
    """Extract transaction ID from order meta_data or stored JSON value"""
    # Check if we have stored value from JSON
    if '_transaction_id' in order:
        return order.get('_transaction_id')

    # Otherwise check meta_data
    meta_data = order.get('meta_data', [])
    if not meta_data:
        return None

    for meta in meta_data:
        if meta.get('key') == '_transactionId':
            return meta.get('value')
    return None

def get_user_name(order: Dict) -> str:
    """Get user name from billing or shipping"""
    billing = order.get('billing', {})
    shipping = order.get('shipping', {})

    # Try billing first
    if billing:
        first_name = billing.get('first_name', '').strip()
        last_name = billing.get('last_name', '').strip()
        if first_name or last_name:
            return f"{first_name} {last_name}".strip()

    # Fallback to shipping
    if shipping:
        first_name = shipping.get('first_name', '').strip()
        last_name = shipping.get('last_name', '').strip()
        if first_name or last_name:
            return f"{first_name} {last_name}".strip()

    return 'نامشخص'

def normalize_name(name: str) -> str:
    """Normalize name for matching (remove extra spaces, convert to lowercase)"""
    return ' '.join(name.split()).lower().strip()

def extract_orders_without_phones(wc_client: WooCommerceClient, start_date: str, end_date: str) -> List[Dict]:
    """Extract orders without phone numbers from WooCommerce API"""
    print("🔍 Step 1: Extracting orders without phone numbers from WooCommerce API...")
    print(f"   Date range: {start_date} to {end_date}")
    print(f"   Status: processing, completed")
    print(f"   Payment: WC_Gateway_SnappPay")
    print()

    all_orders = []
    page = 1
    per_page = 100

    while True:
        print(f"📄 Fetching page {page}...", end=' ')

        params = {
            'after': start_date,
            'before': end_date,
            'payment_method': 'WC_Gateway_SnappPay',
            'status': 'processing,completed',
            'per_page': per_page,
            'page': page,
            'orderby': 'date',
            'order': 'desc'
        }

        try:
            orders, total_pages, total_items = wc_client.get_orders(params)

            if not orders:
                print("No more orders")
                break

            # Filter for orders without phone numbers
            orders_without_phone = [
                order for order in orders
                if order.get('payment_method') == 'WC_Gateway_SnappPay' and
                (not order.get('billing', {}).get('phone') or
                 not order.get('billing', {}).get('phone', '').strip())
            ]

            all_orders.extend(orders_without_phone)
            print(f"Found {len(orders)} orders, {len(orders_without_phone)} without phone numbers")

            if page >= total_pages:
                break

            page += 1
            time.sleep(0.1)  # Reduced rate limiting

        except Exception as e:
            print(f"❌ Error: {e}")
            break

    print(f"\n✅ Total: {len(all_orders)} orders without phone numbers")
    return all_orders

def load_orders_from_json(json_file: Path) -> List[Dict]:
    """Load orders from JSON file for fast lookup"""
    if not json_file.exists():
        return []

    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            orders = json.load(f)
            # Convert to WooCommerce API format for compatibility
            converted_orders = []
            for order in orders:
                # Convert to API format
                converted_order = {
                    'id': order.get('orderId'),
                    'date_created': order.get('orderDate'),
                    'billing': {
                        'first_name': order.get('billing', {}).get('first_name', ''),
                        'last_name': order.get('billing', {}).get('last_name', ''),
                        'phone': order.get('billing', {}).get('phone', ''),
                        'email': order.get('billing', {}).get('email', ''),
                        'city': order.get('billing', {}).get('city', ''),
                    },
                    'shipping': {
                        'first_name': order.get('shipping', {}).get('first_name', ''),
                        'last_name': order.get('shipping', {}).get('last_name', ''),
                        'city': order.get('shipping', {}).get('city', ''),
                    },
                    'total': order.get('total', ''),
                    'status': order.get('status', ''),
                }
                converted_orders.append(converted_order)
            return converted_orders
    except Exception as e:
        print(f"⚠️  Could not load orders from JSON: {e}")
        return []

def search_orders_by_name(wc_clients: List[WooCommerceClient], name: str, start_date: str, end_date: str,
                          cache: Dict = None, json_orders: List[Dict] = None) -> List[Dict]:
    """Search for orders with matching name that have phone numbers - uses JSON file first, then API"""
    # Check cache first
    normalized_name = normalize_name(name)
    if cache is not None and normalized_name in cache:
        cached_result = cache[normalized_name]
        # Return cached result even if empty (to avoid re-searching names with no matches)
        if isinstance(cached_result, list):
            return cached_result

    matching_orders = []

    # First, search in JSON file (much faster!)
    if json_orders:
        for order in json_orders:
            order_name = get_user_name(order)
            phone = order.get('billing', {}).get('phone', '').strip()

            # Check if names match (normalized comparison) and has phone
            order_normalized = normalize_name(order_name)
            if order_normalized == normalized_name and phone:
                matching_orders.append(order)

        # If we found matches in JSON, return them (no need to search API)
        if matching_orders:
            if cache is not None:
                cache[normalized_name] = matching_orders
            return matching_orders

    # Fallback to API search if JSON didn't have matches
    seen_order_ids = set(order.get('id') for order in matching_orders)  # Track IDs from JSON

    # Search across all WooCommerce sources via API using search parameter (like WordPress admin)
    for source_idx, wc_client in enumerate(wc_clients):
        # Use WooCommerce search parameter - searches in billing name, shipping name, etc.
        # This is exactly what WordPress admin does when you search for orders
        search_params = {
            'search': name,  # Search by the full name (WooCommerce will search in billing/shipping names)
            'per_page': 100,
            'page': 1,
            'orderby': 'date',
            'order': 'desc'
            # Note: No 'status' filter - we want to find matches regardless of order status
        }

        try:
            # Search with the name - WooCommerce will return matching orders
            orders, total_pages, _ = wc_client.get_orders(search_params)

            if orders:
                # Filter orders that have phone numbers and verify name match
                for order in orders:
                    order_id = order.get('id')
                    # Skip if we've seen this order from JSON or another source
                    if order_id in seen_order_ids:
                        continue
                    seen_order_ids.add(order_id)

                    order_name = get_user_name(order)
                    phone = order.get('billing', {}).get('phone', '').strip()

                    # Verify name match (normalized comparison) and has phone
                    order_normalized = normalize_name(order_name)
                    if order_normalized == normalized_name and phone:
                        matching_orders.append(order)

                # If there are more pages, search them too (unlikely but possible)
                if total_pages > 1:
                    for page in range(2, total_pages + 1):
                        search_params['page'] = page
                        orders, _, _ = wc_client.get_orders(search_params)

                        if not orders:
                            break

                        for order in orders:
                            order_id = order.get('id')
                            if order_id in seen_order_ids:
                                continue
                            seen_order_ids.add(order_id)

                            order_name = get_user_name(order)
                            phone = order.get('billing', {}).get('phone', '').strip()
                            order_normalized = normalize_name(order_name)
                            if order_normalized == normalized_name and phone:
                                matching_orders.append(order)

                        time.sleep(0.1)  # Reduced rate limiting

        except Exception as e:
            print(f"      ⚠️  Error searching source {source_idx + 1}: {e}")
            continue

    # Cache the result (even if empty, to avoid re-searching)
    if cache is not None:
        cache[normalized_name] = matching_orders

    return matching_orders

def load_name_cache(cache_file: Path) -> Dict:
    """Load cached name search results"""
    if cache_file.exists():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️  Could not load cache: {e}")
    return {}

def save_name_cache(cache_file: Path, cache: Dict):
    """Save name search results to cache"""
    try:
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  Could not save cache: {e}")

def match_phone_numbers(wc_clients: List[WooCommerceClient], orders_without_phone: List[Dict],
                        start_date: str, end_date: str, cache_file: Path = None, json_orders: List[Dict] = None) -> List[Dict]:
    """Match orders without phone numbers to orders with phone numbers by name across multiple sources"""
    print("\n🔍 Step 2: Searching for matching orders with phone numbers...")
    print(f"   Searching across {len(wc_clients)} source(s)...")

    # Load cache if available
    name_cache = {}
    cached_count = 0
    if cache_file and cache_file.exists():
        name_cache = load_name_cache(cache_file)
        # Count only non-empty cached results
        cached_count = sum(1 for v in name_cache.values() if isinstance(v, list) and len(v) > 0)
        if name_cache:
            print(f"   Loaded {len(name_cache)} cached name searches ({cached_count} with results)")
    print()

    guessed_results = []
    unique_names = {}
    cache_updated = False

    try:
        # Group orders by normalized name
        for order in orders_without_phone:
            name = get_user_name(order)
            normalized = normalize_name(name)

            if normalized not in unique_names:
                unique_names[normalized] = {
                    'name': name,
                    'orders': []
                }
            unique_names[normalized]['orders'].append(order)

        print(f"📊 Found {len(unique_names)} unique names to search")
        print()

        for idx, (normalized_name, name_data) in enumerate(unique_names.items(), 1):
            name = name_data['name']
            orders = name_data['orders']

            if name == 'نامشخص':
                continue

            # Check if we have cached result
            if normalized_name in name_cache:
                cached_result = name_cache[normalized_name]
                # Only use cache if it has results (re-search if empty cache)
                if isinstance(cached_result, list) and len(cached_result) > 0:
                    print(f"[{idx}/{len(unique_names)}] Searching for: {name}... (cached)", end=' ')
                    matching_orders = cached_result
                else:
                    # Cache has empty result, but let's re-search to be sure
                    print(f"[{idx}/{len(unique_names)}] Searching for: {name}... (re-searching)", end=' ')
                    matching_orders = search_orders_by_name(wc_clients, name, start_date, end_date, name_cache)
                    cache_updated = True
            else:
                print(f"[{idx}/{len(unique_names)}] Searching for: {name}...", end=' ', flush=True)
                # Search for matching orders across all sources
                matching_orders = search_orders_by_name(wc_clients, name, start_date, end_date, name_cache)
                cache_updated = True

            if matching_orders:
                # Get the most common phone number from matching orders
                phones = {}
                for match_order in matching_orders:
                    phone = match_order.get('billing', {}).get('phone', '').strip()
                    if phone:
                        phones[phone] = phones.get(phone, 0) + 1

                # Get the most frequent phone number
                guessed_phone = max(phones.items(), key=lambda x: x[1])[0] if phones else None

                # Determine confidence based on:
                # 1. Number of matching orders
                # 2. Whether all orders share the same phone number
                total_orders_with_phones = sum(phones.values())
                unique_phone_count = len(phones)

                if total_orders_with_phones == 0:
                    match_confidence = 'low'
                elif unique_phone_count == 1 and total_orders_with_phones >= 2:
                    # All matching orders have the same phone number - highest confidence
                    match_confidence = 'high'
                elif unique_phone_count == 1 and total_orders_with_phones == 1:
                    # Only one order found, but it has a phone number
                    match_confidence = 'medium'
                elif unique_phone_count > 1:
                    # Multiple different phone numbers found - lower confidence
                    # But if most orders share the same number, it's still medium
                    most_common_count = max(phones.values())
                    if most_common_count >= total_orders_with_phones * 0.7:  # 70% or more share the same number
                        match_confidence = 'medium'
                    else:
                        match_confidence = 'low'
                else:
                    match_confidence = 'medium'

                print(f"✅ Found {len(matching_orders)} matching orders ({total_orders_with_phones} with phones, {unique_phone_count} unique), phone: {guessed_phone}, confidence: {match_confidence}")

                # Create result for each order without phone
                for order in orders:
                    snapppay_token = extract_snapppay_token(order)
                    transaction_id = extract_transaction_id(order)

                    if snapppay_token:
                        guessed_results.append({
                            'order_id': order.get('id'),
                            'order_date': order.get('date_created'),
                            'user_name': name,
                            'snapppay_token': snapppay_token,
                            'transaction_id': transaction_id or '',
                            'guessed_phone': guessed_phone or '',
                            'match_confidence': match_confidence,
                            'matching_orders_count': len(matching_orders),
                            'unique_phone_count': unique_phone_count,
                            'billing': {
                                'first_name': order.get('billing', {}).get('first_name', ''),
                                'last_name': order.get('billing', {}).get('last_name', ''),
                                'phone': order.get('billing', {}).get('phone', ''),
                                'email': order.get('billing', {}).get('email', ''),
                                'city': order.get('billing', {}).get('city', ''),
                            },
                            'shipping': {
                                'first_name': order.get('shipping', {}).get('first_name', ''),
                                'last_name': order.get('shipping', {}).get('last_name', ''),
                                'city': order.get('shipping', {}).get('city', ''),
                            },
                            'total': order.get('total', ''),
                            'status': order.get('status', ''),
                        })
            else:
                print("❌ No matches found")

            time.sleep(0.1)  # Reduced rate limiting

    finally:
        # Always save cache if it was updated (even on early exit/interrupt)
        if cache_file and cache_updated:
            save_name_cache(cache_file, name_cache)
            print(f"\n💾 Saved {len(name_cache)} name searches to cache")

    return guessed_results

def write_results(output_file: Path, results: List[Dict], format_type: str = 'json'):
    """Write results to file"""
    if format_type == 'json':
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"\n✅ Written {len(results)} results to JSON: {output_file}")

    elif format_type == 'xlsx' and HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "Guessed Orders"

        # Headers
        headers = ['Order ID', 'Order Date', 'User Name', 'Guessed Phone',
                  'SnappPay Token', 'Transaction ID', 'Match Confidence',
                  'Matching Orders', 'Billing City', 'Total', 'Status']

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        ws.freeze_panes = 'A2'

        # Data
        for row_num, result in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=result['order_id']).border = border
            ws.cell(row=row_num, column=2, value=result['order_date']).border = border
            ws.cell(row=row_num, column=3, value=result['user_name']).border = border
            ws.cell(row=row_num, column=4, value=result['guessed_phone']).border = border
            ws.cell(row=row_num, column=5, value=result['snapppay_token']).border = border
            ws.cell(row=row_num, column=6, value=result['transaction_id']).border = border
            ws.cell(row=row_num, column=7, value=result['match_confidence']).border = border
            ws.cell(row=row_num, column=8, value=result['matching_orders_count']).border = border
            ws.cell(row=row_num, column=9, value=result['billing']['city']).border = border
            ws.cell(row=row_num, column=10, value=result['total']).border = border
            ws.cell(row=row_num, column=11, value=result['status']).border = border

            if row_num % 2 == 0:
                fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).fill = fill

        # Auto-adjust columns
        column_widths = [12, 20, 25, 15, 40, 25, 15, 15, 15, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.row_dimensions[1].height = 30
        wb.save(output_file)
        print(f"\n✅ Written {len(results)} results to Excel: {output_file}")

    elif format_type == 'html':
        # WooCommerce API credentials (embedded in HTML for internal use)
        wc_base_url = 'https://infinitycolor.co'
        wc_consumer_key = 'WOOCOMMERCE_CONSUMER_KEY'
        wc_consumer_secret = 'WOOCOMMERCE_CONSUMER_SECRET'

        # Serialize orders data for JavaScript (use base64 encoding to avoid escaping issues)
        orders_json = json.dumps(results, ensure_ascii=False)
        orders_json_b64 = base64.b64encode(orders_json.encode('utf-8')).decode('ascii')

        html_content = f"""<!DOCTYPE html>
<html lang="fa" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Guessed Orders with Phone Numbers</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}

        .container {{
            max-width: 1600px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}

        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}

        .stats {{
            display: flex;
            justify-content: space-around;
            padding: 20px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
        }}

        .stat-item {{
            text-align: center;
        }}

        .stat-value {{
            font-size: 32px;
            font-weight: bold;
            color: #667eea;
        }}

        .stat-label {{
            font-size: 14px;
            color: #6c757d;
            margin-top: 5px;
        }}

        .table-wrapper {{
            overflow-x: auto;
            padding: 20px;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}

        thead {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }}

        th {{
            padding: 15px;
            text-align: right;
            font-weight: 600;
            border: 1px solid rgba(255,255,255,0.2);
        }}

        td {{
            padding: 12px 15px;
            border: 1px solid #e9ecef;
        }}

        tbody tr {{
            transition: background-color 0.2s;
        }}

        tbody tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}

        tbody tr:hover {{
            background-color: #e7f3ff;
            cursor: pointer;
        }}

        .order-id {{
            font-weight: 600;
            color: #667eea;
        }}

        .phone {{
            font-weight: 600;
            color: #28a745;
        }}

        .token {{
            font-family: 'Courier New', monospace;
            font-size: 12px;
            color: #495057;
            word-break: break-all;
        }}

        .confidence-high {{
            background-color: #d4edda;
            color: #155724;
            padding: 4px 8px;
            border-radius: 4px;
            font-weight: 600;
        }}

        .confidence-medium {{
            background-color: #fff3cd;
            color: #856404;
            padding: 4px 8px;
            border-radius: 4px;
            font-weight: 600;
        }}

        .footer {{
            padding: 20px;
            text-align: center;
            color: #6c757d;
            font-size: 12px;
            border-top: 2px solid #e9ecef;
            background: #f8f9fa;
        }}

        .btn-update {{
            background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-weight: 600;
            font-size: 13px;
            transition: all 0.3s;
        }}

        .btn-update:hover:not(:disabled) {{
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(40, 167, 69, 0.3);
        }}

        .btn-update:disabled {{
            background: #6c757d;
            cursor: not-allowed;
            opacity: 0.6;
        }}

        .btn-update.loading {{
            background: #ffc107;
            position: relative;
        }}

        .btn-update.loading::after {{
            content: '';
            position: absolute;
            width: 16px;
            height: 16px;
            top: 50%;
            left: 50%;
            margin-left: -8px;
            margin-top: -8px;
            border: 2px solid #ffffff;
            border-radius: 50%;
            border-top-color: transparent;
            animation: spin 0.8s linear infinite;
        }}

        @keyframes spin {{
            to {{ transform: rotate(360deg); }}
        }}

        .btn-bulk {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 12px 24px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 600;
            font-size: 14px;
            margin: 10px;
            transition: all 0.3s;
        }}

        .btn-bulk:hover:not(:disabled) {{
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
        }}

        .btn-bulk:disabled {{
            background: #6c757d;
            cursor: not-allowed;
            opacity: 0.6;
        }}

        .status-updated {{
            background-color: #d4edda;
            color: #155724;
            padding: 4px 8px;
            border-radius: 4px;
            font-weight: 600;
            font-size: 12px;
        }}

        .status-error {{
            background-color: #f8d7da;
            color: #721c24;
            padding: 4px 8px;
            border-radius: 4px;
            font-weight: 600;
            font-size: 12px;
        }}

        .message-area {{
            padding: 15px;
            margin: 10px 20px;
            border-radius: 8px;
            display: none;
        }}

        .message-success {{
            background-color: #d4edda;
            color: #155724;
            border: 1px solid #c3e6cb;
        }}

        .message-error {{
            background-color: #f8d7da;
            color: #721c24;
            border: 1px solid #f5c6cb;
        }}

        .bulk-controls {{
            padding: 20px;
            background: #f8f9fa;
            border-bottom: 2px solid #e9ecef;
            display: flex;
            align-items: center;
            justify-content: space-between;
            flex-wrap: wrap;
        }}

        .bulk-info {{
            display: flex;
            align-items: center;
            gap: 15px;
        }}

        .checkbox-bulk {{
            width: 18px;
            height: 18px;
            cursor: pointer;
        }}
    </style>
    <script>
        // WooCommerce API Configuration
        const WC_CONFIG = {{
            baseUrl: '{wc_base_url}',
            consumerKey: '{wc_consumer_key}',
            consumerSecret: '{wc_consumer_secret}'
        }};

        // Orders data embedded in HTML (base64 encoded for safety)
        const ORDERS_DATA_JSON = atob('{orders_json_b64}');
        const ORDERS_DATA = JSON.parse(ORDERS_DATA_JSON);

        // Update single order
        async function updateOrderPhone(orderId, phoneNumber, buttonElement) {{
            if (!phoneNumber || phoneNumber === 'N/A') {{
                showMessage('خطا: شماره تلفن معتبر نیست', 'error');
                return;
            }}

            const row = buttonElement.closest('tr');
            const statusCell = row.querySelector('.update-status');

            // Disable button and show loading
            buttonElement.disabled = true;
            buttonElement.classList.add('loading');
            buttonElement.textContent = '';

            try {{
                const url = `${{WC_CONFIG.baseUrl}}/wp-json/wc/v3/orders/${{orderId}}`;
                const params = new URLSearchParams({{
                    consumer_key: WC_CONFIG.consumerKey,
                    consumer_secret: WC_CONFIG.consumerSecret
                }});

                const response = await fetch(`${{url}}?${{params}}`, {{
                    method: 'PUT',
                    headers: {{
                        'Content-Type': 'application/json',
                    }},
                    body: JSON.stringify({{
                        billing: {{
                            phone: phoneNumber
                        }}
                    }})
                }});

                if (!response.ok) {{
                    const errorData = await response.json().catch(() => ({{}}));
                    throw new Error(errorData.message || `HTTP ${{response.status}}: ${{response.statusText}}`);
                }}

                const updatedOrder = await response.json();

                // Success
                buttonElement.disabled = true;
                buttonElement.classList.remove('loading');
                buttonElement.textContent = '✓ به‌روزرسانی شد';
                buttonElement.style.background = '#28a745';

                if (statusCell) {{
                    statusCell.innerHTML = '<span class="status-updated">به‌روزرسانی شد</span>';
                }}

                showMessage(`سفارش #${{orderId}} با موفقیت به‌روزرسانی شد`, 'success');

                // Update stats
                updateStats();

            }} catch (error) {{
                console.error('Update error:', error);
                buttonElement.disabled = false;
                buttonElement.classList.remove('loading');
                buttonElement.textContent = 'به‌روزرسانی';

                if (statusCell) {{
                    statusCell.innerHTML = `<span class="status-error">خطا: ${{error.message}}</span>`;
                }}

                showMessage(`خطا در به‌روزرسانی سفارش #${{orderId}}: ${{error.message}}`, 'error');
            }}
        }}

        // Bulk update high confidence orders
        async function updateBulkOrders() {{
            const highConfidenceOrders = ORDERS_DATA.filter(o =>
                o.match_confidence === 'high' &&
                o.guessed_phone &&
                o.guessed_phone !== 'N/A'
            );

            if (highConfidenceOrders.length === 0) {{
                showMessage('هیچ سفارشی با اطمینان بالا برای به‌روزرسانی وجود ندارد', 'error');
                return;
            }}

            if (!confirm(`آیا می‌خواهید ${{highConfidenceOrders.length}} سفارش با اطمینان بالا را به‌روزرسانی کنید؟`)) {{
                return;
            }}

            const bulkButton = document.getElementById('btn-bulk-update');
            bulkButton.disabled = true;
            bulkButton.textContent = `در حال به‌روزرسانی... (0/${{highConfidenceOrders.length}})`;

            let successCount = 0;
            let errorCount = 0;

            for (let i = 0; i < highConfidenceOrders.length; i++) {{
                const order = highConfidenceOrders[i];
                const buttonElement = document.querySelector(`button[data-order-id="${{order.order_id}}"]`);

                if (buttonElement && !buttonElement.disabled) {{
                    await updateOrderPhone(order.order_id, order.guessed_phone, buttonElement);

                    if (buttonElement.disabled && buttonElement.textContent.includes('✓')) {{
                        successCount++;
                    }} else {{
                        errorCount++;
                    }}

                    // Update progress
                    bulkButton.textContent = `در حال به‌روزرسانی... (${{i + 1}}/${{highConfidenceOrders.length}})`;

                    // Small delay to avoid overwhelming the API
                    await new Promise(resolve => setTimeout(resolve, 500));
                }}
            }}

            bulkButton.disabled = false;
            bulkButton.textContent = 'به‌روزرسانی دسته‌ای (اطمینان بالا)';

            showMessage(
                `به‌روزرسانی کامل شد: ${{successCount}} موفق، ${{errorCount}} خطا`,
                successCount > 0 ? 'success' : 'error'
            );
        }}

        // Show message
        function showMessage(message, type) {{
            const messageArea = document.getElementById('message-area');
            messageArea.textContent = message;
            messageArea.className = `message-area message-${{type}}`;
            messageArea.style.display = 'block';

            setTimeout(() => {{
                messageArea.style.display = 'none';
            }}, 5000);
        }}

        // Update statistics
        function updateStats() {{
            const updatedCount = document.querySelectorAll('.btn-update:disabled').length;
            const totalCount = ORDERS_DATA.length;
            const updatedStat = document.getElementById('stat-updated');
            if (updatedStat) {{
                updatedStat.textContent = updatedCount;
            }}
        }}

        // Initialize on page load
        document.addEventListener('DOMContentLoaded', function() {{
            updateStats();
        }});
    </script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Guessed Orders with Phone Numbers</h1>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>

        <div class="stats">
            <div class="stat-item">
                <div class="stat-value">{len(results)}</div>
                <div class="stat-label">Total Orders</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(1 for r in results if r['guessed_phone'])}</div>
                <div class="stat-label">With Guessed Phones</div>
            </div>
            <div class="stat-item">
                <div class="stat-value">{sum(1 for r in results if r['match_confidence'] == 'high')}</div>
                <div class="stat-label">High Confidence</div>
            </div>
            <div class="stat-item">
                <div class="stat-value" id="stat-updated">0</div>
                <div class="stat-label">Updated</div>
            </div>
        </div>

        <div id="message-area" class="message-area"></div>

        <div class="bulk-controls">
            <div class="bulk-info">
                <button id="btn-bulk-update" class="btn-bulk" onclick="updateBulkOrders()">
                    به‌روزرسانی دسته‌ای (اطمینان بالا)
                </button>
                <span style="color: #6c757d; font-size: 14px;">
                    {sum(1 for r in results if r['match_confidence'] == 'high' and r['guessed_phone'])} سفارش با اطمینان بالا
                </span>
            </div>
        </div>

        <div class="table-wrapper">
            <table>
                <thead>
                    <tr>
                        <th>Order ID</th>
                        <th>Order Date</th>
                        <th>User Name</th>
                        <th>Guessed Phone</th>
                        <th>SnappPay Token</th>
                        <th>Transaction ID</th>
                        <th>Confidence</th>
                        <th>Matching Orders</th>
                        <th>City</th>
                        <th>Total</th>
                        <th>Status</th>
                        <th>Action</th>
                        <th>Update Status</th>
                    </tr>
                </thead>
                <tbody>
"""

        for result in results:
            confidence_class = f"confidence-{result['match_confidence']}"
            phone_value = result['guessed_phone'] or 'N/A'
            has_phone = phone_value != 'N/A'
            button_disabled = '' if has_phone else 'disabled'

            html_content += f"""
                    <tr data-order-id="{result['order_id']}">
                        <td class="order-id">{result['order_id']}</td>
                        <td>{result['order_date']}</td>
                        <td>{result['user_name']}</td>
                        <td class="phone">{phone_value}</td>
                        <td class="token">{result['snapppay_token']}</td>
                        <td class="token">{result['transaction_id'] or ''}</td>
                        <td><span class="{confidence_class}">{result['match_confidence'].upper()}</span></td>
                        <td>{result['matching_orders_count']}</td>
                        <td>{result['billing']['city']}</td>
                        <td>{result['total']}</td>
                        <td>{result['status']}</td>
                        <td>
                            <button
                                class="btn-update"
                                data-order-id="{result['order_id']}"
                                onclick="updateOrderPhone({result['order_id']}, '{phone_value}', this)"
                                {button_disabled}>
                                به‌روزرسانی
                            </button>
                        </td>
                        <td class="update-status"></td>
                    </tr>
"""

        html_content += """
                </tbody>
            </table>
        </div>

        <div class="footer">
            <p>Guessed Orders Data - Infinity Store | Click "به‌روزرسانی" to update phone numbers in WooCommerce</p>
        </div>
    </div>
</body>
</html>
"""

        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"\n✅ Written {len(results)} results to HTML: {output_file}")

def main():
    import sys

    # Check for --clear-cache flag
    clear_cache = '--clear-cache' in sys.argv or '-c' in sys.argv

    # WooCommerce configuration (from woocommerce-importer/config.js)
    # Primary source
    WC_BASE_URL = 'https://infinitycolor.co'
    WC_CONSUMER_KEY = 'WOOCOMMERCE_CONSUMER_KEY'
    WC_CONSUMER_SECRET = 'WOOCOMMERCE_CONSUMER_SECRET'

    # Backup sources for matching (add your backup URLs here)
    # Format: (base_url, consumer_key, consumer_secret)
    # These will be used ONLY for searching matches, not for extraction
    BACKUP_SOURCES = [
        ('https://pasdaranbookcity.ir/', 'ck_ab4e8240c3b7d9c38b4a557a97124d4450d497ff', 'cs_b23949f478cc40e25d23cce261d20da0e44f6199'),
    ]

    # Date range
    start_date = '2025-11-27T00:00:00'
    end_date = datetime.now().isoformat()

    # Output file
    script_dir = Path(__file__).parent
    output_file = script_dir / 'guessed-orders-with-phones.json'
    cache_file = script_dir / 'name-search-cache.json'
    json_orders_file = script_dir / 'woocommerce-guest-orders-snapppay-data.json'

    # Clear cache if requested
    if clear_cache and cache_file.exists():
        cache_file.unlink()
        print("🗑️  Cache cleared")
        print()

    # Load orders from JSON file for fast name matching (optional, for faster matching)
    print("📂 Loading orders from JSON file for fast name matching...", end=' ')
    json_orders = load_orders_from_json(json_orders_file)
    if json_orders:
        print(f"✅ Loaded {len(json_orders)} orders for name matching")
    else:
        print("⚠️  JSON file not found or empty, will use API only for matching")
    print()

    print("🚀 WooCommerce Order Phone Number Matcher")
    print("=" * 60)
    print(f"📅 Date range: {start_date} to {end_date}")
    print(f"🌐 Primary WooCommerce: {WC_BASE_URL}")
    print()

    # Initialize primary WooCommerce client (for extraction)
    primary_client = WooCommerceClient(WC_BASE_URL, WC_CONSUMER_KEY, WC_CONSUMER_SECRET)

    # Initialize backup clients for matching
    matching_clients = [primary_client]  # Always include primary
    for backup_url, backup_key, backup_secret in BACKUP_SOURCES:
        try:
            backup_client = WooCommerceClient(backup_url, backup_key, backup_secret)
            matching_clients.append(backup_client)
            print(f"✅ Added backup source: {backup_url}")
        except Exception as e:
            print(f"⚠️  Failed to initialize backup source {backup_url}: {e}")

    if len(matching_clients) > 1:
        print(f"📊 Will search across {len(matching_clients)} sources for phone number matches")
    print()

    try:
        # Step 1: Extract orders without phone numbers from WooCommerce API
        orders_without_phone = extract_orders_without_phones(primary_client, start_date, end_date)

        if not orders_without_phone:
            print("\n⚠️  No orders without phone numbers found")
            return

        # Step 2: Match phone numbers by searching for orders with same names (search JSON first, then API)
        guessed_results = match_phone_numbers(matching_clients, orders_without_phone, start_date, end_date, cache_file, json_orders)

        if not guessed_results:
            print("\n⚠️  No phone number matches found")
            return

        # Step 3: Write results
        write_results(output_file, guessed_results, 'json')

        # Also write Excel and HTML if available
        if HAS_OPENPYXL:
            excel_file = script_dir / 'guessed-orders-with-phones.xlsx'
            write_results(excel_file, guessed_results, 'xlsx')

        html_file = script_dir / 'guessed-orders-with-phones.html'
        write_results(html_file, guessed_results, 'html')

        # Print summary
        print("\n📊 Summary:")
        print(f"   Orders without phones: {len(orders_without_phone)}")
        print(f"   Orders with guessed phones: {len(guessed_results)}")
        print(f"   High confidence matches: {sum(1 for r in guessed_results if r['match_confidence'] == 'high')}")
        print(f"   Medium confidence matches: {sum(1 for r in guessed_results if r['match_confidence'] == 'medium')}")
        print(f"   Orders with phone numbers: {sum(1 for r in guessed_results if r['guessed_phone'])}")

        # Print sample
        if guessed_results:
            print("\n📋 Sample result:")
            sample = guessed_results[0]
            print(f"   Order ID: {sample['order_id']}")
            print(f"   User Name: {sample['user_name']}")
            print(f"   Guessed Phone: {sample['guessed_phone']}")
            print(f"   Confidence: {sample['match_confidence']}")
            print(f"   Matching Orders: {sample['matching_orders_count']}")

    except KeyboardInterrupt:
        print("\n\n⚠️  Interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()

